import openpyxl
import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions



path ="C:\\Users\\admin\\Downloads\\download.xlsx"
Dict={}

driver =webdriver.Chrome()
driver.maximize_window()
driver.implicitly_wait(5)


driver.get("https://rahulshettyacademy.com/upload-download-test/index.html")

price=int(driver.find_element(By.XPATH,"//div[text()='Apple']/parent::div/parent::div/div[@id='cell-4-undefined']").text) 
print(price) 
driver.find_element(By.CSS_SELECTOR,"#downloadButton").click()
time.sleep(3)

book=openpyxl.load_workbook(path)
sheet=book.active
for i in range(1,sheet.max_column+1):
    if sheet.cell(row=1,column=i).value =="price":
        for j in range(i,sheet.max_column+1):
            Dict[sheet.cell(row=1,column=j).value]=sheet.cell(row=i,column=j).value

print(Dict)
locl = driver.find_element(By.CSS_SELECTOR,"input[type='file']")
locl.send_keys(path)

wait=WebDriverWait(driver,5)
toast_locator = (By.CSS_SELECTOR,".Toastify__toast-body div:nth-child(2)")
wait.until(expected_conditions.visibility_of_element_located(toast_locator))

print(driver.find_element(*toast_locator).text)
